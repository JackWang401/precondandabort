import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from precond_abort.analyzer import AbortAnalyzer, combine_analysis_results
from precond_abort.calibration import CalibrationRepository
from precond_abort.mapping import default_mapping
from precond_abort.report import OUTPUT_HEADER_ROW_1, OUTPUT_HEADER_ROW_2, write_report
import tests.test_analyzer as analyzer_fixtures


class ReportTests(unittest.TestCase):
    def test_report_contains_requested_and_audit_sheets(self):
        fixture = analyzer_fixtures.AbortAnalyzerTests()
        fixture.setUp()
        result = AbortAnalyzer().analyze(
            fixture._source(), fixture.mapping, fixture.calibrations, "input.mf4"
        )
        with tempfile.TemporaryDirectory() as directory:
            path = write_report(result, Path(directory) / "report.xlsx")
            workbook = load_workbook(path, data_only=False)
            self.assertEqual(
                workbook.sheetnames,
                ["Abort Analysis", "Event Details", "Signal Mapping", "Parameters", "Run Information"],
            )
            analysis = workbook["Abort Analysis"]
            header_row_1 = [cell.value for cell in analysis[1]]
            header_row_2 = [cell.value for cell in analysis[2]]
            self.assertEqual(header_row_1, list(OUTPUT_HEADER_ROW_1))
            self.assertEqual(header_row_2, list(OUTPUT_HEADER_ROW_2))
            self.assertEqual(
                {str(cell_range) for cell_range in analysis.merged_cells.ranges},
                {"A1:A2", "B1:B2", "C1:C2", "D1:F1", "G1:I1", "J1:L1", "M1:O1", "P1:R1", "S1:U1", "V1:V2"},
            )
            self.assertEqual(analysis.max_row, 5)
            self.assertEqual(analysis["A3"].value, "input.mf4")
            self.assertEqual(analysis["B3"].value, 1)
            self.assertEqual(analysis["C3"].value, 10)
            self.assertEqual(analysis["D3"].value, 11)
            self.assertEqual(analysis["E3"].value, 10)
            self.assertEqual(analysis["F3"].value, "Yes")
            self.assertEqual(analysis["V5"].value, "Yes")
            for checked_cell in (analysis["F3"], analysis["V5"]):
                self.assertEqual(checked_cell.fill.fill_type, "solid")
                self.assertEqual(checked_cell.fill.fgColor.rgb, "00C6EFCE")
                self.assertTrue(checked_cell.font.bold)
                self.assertEqual(checked_cell.font.color.rgb, "00006100")
            parameter_headers = [cell.value for cell in workbook["Parameters"][1]]
            self.assertEqual(parameter_headers[2:4], ["X Source", "Y Source"])
            self.assertEqual(workbook["Run Information"]["B5"].value, 3)
            self.assertEqual(workbook["Run Information"]["B7"].value, "Disabled")
            workbook.close()

    def test_report_records_enabled_throttle_mode_and_values(self):
        fixture = analyzer_fixtures.AbortAnalyzerTests()
        fixture.setUp()
        result = AbortAnalyzer().analyze(
            fixture._source(),
            fixture.mapping,
            fixture.calibrations,
            "input.mf4",
            enable_throttle_checks=True,
        )
        with tempfile.TemporaryDirectory() as directory:
            path = write_report(result, Path(directory) / "report.xlsx")
            workbook = load_workbook(path, data_only=False)
            analysis = workbook["Abort Analysis"]
            self.assertEqual(analysis["P4"].value, 60)
            self.assertEqual(analysis["Q4"].value, 20)
            self.assertEqual(analysis["R4"].value, "Yes")
            self.assertEqual(analysis["S4"].value, 70)
            self.assertEqual(analysis["T4"].value, 85)
            self.assertEqual(analysis["U4"].value, "No")
            self.assertEqual(workbook["Run Information"]["B7"].value, "Enabled")
            workbook.close()

    def test_report_lists_every_input_measurement(self):
        fixture = analyzer_fixtures.AbortAnalyzerTests()
        fixture.setUp()
        results = tuple(
            AbortAnalyzer().analyze(
                fixture._source(), fixture.mapping, fixture.calibrations, filename
            )
            for filename in ("first.mf4", "second.mdf")
        )
        with tempfile.TemporaryDirectory() as directory:
            path = write_report(
                combine_analysis_results(results),
                Path(directory) / "batch_report.xlsx",
            )
            workbook = load_workbook(path, data_only=False)
            self.assertEqual(
                workbook["Run Information"]["B2"].value,
                "first.mf4\nsecond.mdf",
            )
            self.assertEqual(workbook["Abort Analysis"]["A6"].value, "second.mdf")
            self.assertEqual(workbook["Run Information"]["B5"].value, 6)
            workbook.close()


if __name__ == "__main__":
    unittest.main()
