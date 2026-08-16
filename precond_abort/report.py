from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .errors import InputValidationError
from .models import AnalysisResult


OUTPUT_GROUPS = (
    ("steering wheel angle", "strAng"),
    ("steering wheel angle speed", "strAngSpd"),
    ("yaw rate", "yawRate"),
    ("lateral acceleration", "latAccel"),
    ("throttle increase", "throttleInc"),
    ("maximum throttle", "maxThrottle"),
)

OUTPUT_REASON_FLAGS = tuple(flag for _, flag in OUTPUT_GROUPS) + ("others",)

OUTPUT_HEADER_ROW_1 = (
    "File Name",
    "timestamp",
    "speed",
    "steering wheel angle",
    None,
    None,
    "steering wheel angle speed",
    None,
    None,
    "yaw rate",
    None,
    None,
    "lateral acceleration",
    None,
    None,
    "throttle increase",
    None,
    None,
    "maximum throttle",
    None,
    None,
    "others",
)

OUTPUT_HEADER_ROW_2 = (
    None,
    None,
    None,
    "actual",
    "thd",
    "result",
    "actual",
    "thd",
    "result",
    "actual",
    "thd",
    "result",
    "actual",
    "thd",
    "result",
    "actual",
    "thd",
    "result",
    "actual",
    "thd",
    "result",
    None,
)

OUTPUT_PREVIEW_HEADERS = (
    "timestamp",
    "speed",
    "angle actual",
    "angle thd",
    "angle result",
    "angle speed actual",
    "angle speed thd",
    "angle speed result",
    "yaw actual",
    "yaw thd",
    "yaw result",
    "lat accel actual",
    "lat accel thd",
    "lat accel result",
    "throttle inc actual",
    "throttle inc thd",
    "throttle inc result",
    "max throttle actual",
    "max throttle thd",
    "max throttle result",
    "others",
)

DETAIL_HEADERS = (
    "filename",
    "timestamp",
    "reasons",
    "vehicleSpeed",
    "steeringWheelAngle",
    "steeringWheelAngle_thd",
    "steeringWheelAngleRate",
    "steeringWheelAngleRate_thd",
    "yawRate",
    "yawRate_thd",
    "lateralAcceleration",
    "lateralAcceleration_thd",
    "throttle",
    "throttleBaseline",
    "throttleIncrease",
    "throttleIncrease_thd",
    "throttleOverride_thd",
    "maxThrottle_thd",
    "decelerationStart",
    "aebDecelerationRequest",
)

NAVY = "17324D"
TEAL = "197C83"
LIGHT_TEAL = "DCEFF0"
LIGHT_BLUE = "E8F0F7"
GREEN = "C6EFCE"
GREEN_TEXT = "006100"
AMBER = "FCE8C3"
GREY = "EDF0F2"
WHITE = "FFFFFF"
TEXT = "24333F"


def write_report(result: AnalysisResult, output_path: str | Path) -> Path:
    output = Path(output_path)
    if output.suffix.lower() != ".xlsx":
        raise InputValidationError(f"Output report must use the .xlsx extension: {output}")
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    analysis = workbook.active
    analysis.title = "Abort Analysis"
    analysis.append(list(OUTPUT_HEADER_ROW_1))
    analysis.append(list(OUTPUT_HEADER_ROW_2))
    for event in result.events:
        analysis.append(event.output_row())
    _format_analysis_sheet(analysis)

    details = workbook.create_sheet("Event Details")
    details.append(list(DETAIL_HEADERS))
    for event in result.events:
        values = event.signal_values
        thresholds = event.thresholds
        details.append(
            [
                event.filename,
                event.timestamp,
                ", ".join(event.reasons),
                event.vehicle_speed,
                values["steering_wheel_angle"],
                thresholds["steering_wheel_angle"],
                values["steering_wheel_angle_rate"],
                thresholds["steering_wheel_angle_rate"],
                values["yaw_rate"],
                thresholds["yaw_rate"],
                values["lateral_acceleration"],
                thresholds["lateral_acceleration"],
                values.get("throttle"),
                event.throttle_baseline,
                event.throttle_increase,
                thresholds.get("throttle_increase"),
                thresholds.get("throttle_override"),
                thresholds.get("throttle_max"),
                event.deceleration_start,
                values.get("aeb_deceleration_request"),
            ]
        )
    _format_table_sheet(details, "EventDetailsTable")
    _set_widths(details, [108, 16, 34] + [22] * (len(DETAIL_HEADERS) - 3))
    for row in details.iter_rows(min_row=2, min_col=2, max_col=20):
        for cell in row:
            cell.number_format = "0.000000"

    mapping_sheet = workbook.create_sheet("Signal Mapping")
    mapping_sheet.append(["Logical Name", "modelLogger", "cal_thd"])
    for logical_name, row in result.mapping.signals.items():
        mapping_sheet.append([logical_name, row.model_logger, "; ".join(row.calibrations)])
    _format_table_sheet(mapping_sheet, "SignalMappingTable")
    _set_widths(mapping_sheet, [34, 78, 42])

    parameters_sheet = workbook.create_sheet("Parameters")
    parameters_sheet.append(
        [
            "Requested Name",
            "Resolved Source",
            "X Source",
            "Y Source",
            "X Unit",
            "Y Unit",
            "X Values",
            "Y Values",
        ]
    )
    for requested_name, parameter in result.parameters.items():
        parameters_sheet.append(
            [
                requested_name,
                parameter.source,
                parameter.x_source or "Constant — no X axis",
                parameter.y_source,
                parameter.x_unit,
                parameter.y_unit,
                _join_numbers(parameter.x_values),
                _join_numbers(parameter.y_values),
            ]
        )
    _format_table_sheet(parameters_sheet, "ParametersTable")
    _set_widths(parameters_sheet, [36, 68, 68, 68, 14, 14, 74, 74])
    for row in parameters_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    metadata = workbook.create_sheet("Run Information")
    metadata_rows = [
        ("Precondition and Abort Analysis", ""),
        ("Input measurement", str(result.input_file)),
        ("Mapping source", result.mapping.source),
        ("Generated (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")),
        ("Abort events", len(result.events)),
        ("Warnings", "\n".join(result.warnings) if result.warnings else "None"),
    ]
    for row in metadata_rows:
        metadata.append(list(row))
    metadata.sheet_view.showGridLines = False
    metadata.merge_cells("A1:B1")
    metadata["A1"].fill = PatternFill("solid", fgColor=NAVY)
    metadata["A1"].font = Font(name="Aptos Display", size=16, bold=True, color=WHITE)
    metadata["A1"].alignment = Alignment(vertical="center")
    metadata.row_dimensions[1].height = 32
    for row in range(2, metadata.max_row + 1):
        metadata.cell(row, 1).font = Font(name="Aptos", bold=True, color=TEXT)
        metadata.cell(row, 1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        metadata.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
    metadata.column_dimensions["A"].width = 24
    metadata.column_dimensions["B"].width = 92
    metadata.row_dimensions[6].height = max(28, 16 * (len(result.warnings) + 1))

    temporary = output.with_name(f".{output.stem}.tmp.xlsx")
    try:
        workbook.save(temporary)
        temporary.replace(output)
    except OSError as exc:
        if temporary.exists():
            temporary.unlink()
        raise InputValidationError(f"Cannot write report {output}: {exc}") from exc
    finally:
        workbook.close()
    return output


def _format_table_sheet(sheet, table_name: str, reason_columns=()) -> None:
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 38
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Aptos", size=11, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if sheet.max_row > 1:
        table = Table(displayName=table_name, ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    thin = Side(style="thin", color="D9E1E6")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Aptos", size=10, color=TEXT)
            cell.alignment = Alignment(vertical="center")
            cell.border = Border(bottom=thin)
    for column_index in reason_columns:
        if sheet.max_row < 2:
            continue
        for row_index in range(2, sheet.max_row + 1):
            sheet.cell(row_index, column_index).alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
        range_ref = f"{sheet.cell(2, column_index).coordinate}:{sheet.cell(sheet.max_row, column_index).coordinate}"
        sheet.conditional_formatting.add(
            range_ref,
            CellIsRule(
                operator="equal",
                formula=['"Yes"'],
                fill=PatternFill("solid", fgColor=GREEN),
                font=Font(bold=True, color=GREEN_TEXT),
            ),
        )
        sheet.conditional_formatting.add(
            range_ref,
            CellIsRule(operator="equal", formula=['"No"'], fill=PatternFill("solid", fgColor=GREY)),
        )


def _format_analysis_sheet(sheet) -> None:
    sheet.freeze_panes = "D3"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0

    header_fill = PatternFill("solid", fgColor=NAVY)
    subheader_fill = PatternFill("solid", fgColor=TEAL)
    header_font = Font(name="Aptos", size=11, bold=True, color=WHITE)
    thin = Side(style="thin", color="D9E1E6")
    medium = Side(style="medium", color="B9C7D0")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    for cell in sheet[2]:
        cell.fill = subheader_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=medium)

    sheet.merge_cells("A1:A2")
    sheet.merge_cells("B1:B2")
    sheet.merge_cells("C1:C2")
    sheet.merge_cells("D1:F1")
    sheet.merge_cells("G1:I1")
    sheet.merge_cells("J1:L1")
    sheet.merge_cells("M1:O1")
    sheet.merge_cells("P1:R1")
    sheet.merge_cells("S1:U1")
    sheet.merge_cells("V1:V2")
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 25

    for row_index in range(3, sheet.max_row + 1):
        row_fill = PatternFill("solid", fgColor=LIGHT_BLUE if row_index % 2 == 0 else WHITE)
        for column_index in range(1, 23):
            cell = sheet.cell(row_index, column_index)
            cell.font = Font(name="Aptos", size=10, color=TEXT)
            cell.fill = row_fill
            cell.alignment = Alignment(
                horizontal="left" if column_index == 1 else "right",
                vertical="center",
            )
            right_border = medium if column_index in {3, 6, 9, 12, 15, 18, 21, 22} else thin
            cell.border = Border(bottom=thin, right=right_border)

    result_columns = (6, 9, 12, 15, 18, 21, 22)
    if sheet.max_row >= 3:
        for column_index in result_columns:
            for row_index in range(3, sheet.max_row + 1):
                cell = sheet.cell(row_index, column_index)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )
                if cell.value == "Yes":
                    cell.fill = PatternFill("solid", fgColor=GREEN)
                    cell.font = Font(
                        name="Aptos",
                        size=10,
                        bold=True,
                        color=GREEN_TEXT,
                    )
            range_ref = (
                f"{sheet.cell(3, column_index).coordinate}:"
                f"{sheet.cell(sheet.max_row, column_index).coordinate}"
            )
            sheet.conditional_formatting.add(
                range_ref,
                CellIsRule(
                    operator="equal",
                    formula=['"Yes"'],
                    fill=PatternFill("solid", fgColor=GREEN),
                    font=Font(bold=True, color=GREEN_TEXT),
                ),
            )
            sheet.conditional_formatting.add(
                range_ref,
                CellIsRule(
                    operator="equal",
                    formula=['"No"'],
                    fill=PatternFill("solid", fgColor=GREY),
                ),
            )

    _set_widths(
        sheet,
        [108, 16, 12]
        + [15, 15, 12] * len(OUTPUT_GROUPS)
        + [12],
    )
    for cell in sheet["B"][2:]:
        cell.number_format = "0.000000"
    numeric_columns = (3, 4, 5, 7, 8, 10, 11, 13, 14, 16, 17, 19, 20)
    for column_index in numeric_columns:
        for row_index in range(3, sheet.max_row + 1):
            sheet.cell(row_index, column_index).number_format = "0.000"


def _set_widths(sheet, widths: list[float]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _join_numbers(values: tuple[float, ...]) -> str:
    return ", ".join(f"{value:g}" for value in values)
