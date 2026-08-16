from __future__ import annotations

import re
import warnings
from pathlib import Path

from openpyxl import Workbook, load_workbook

from .errors import MappingError
from .models import CalibrationBindingSpec, MappingConfiguration, SignalMapping


DEFAULT_SIGNAL_ROWS = (
    SignalMapping("vehicle_speed", "rov/vehicleSpeed"),
    SignalMapping("steering_wheel_angle", "rov/SteeringWheelAngle", ("SteeringWheelAngle_Th",)),
    SignalMapping(
        "steering_wheel_angle_rate",
        "rov/SteeringWheelAngleRate",
        ("AEB_SteeringAngleRate_Override",),
    ),
    SignalMapping("yaw_rate", "rov/YawrateSuspension", ("YawrateSuspension_Th",)),
    SignalMapping("lateral_acceleration", "rov/lateralAcceleration", ("LateralAcceleration_th",)),
    SignalMapping(
        "throttle",
        "rov/PedalPosPro",
        (
            "PedalPosProIncrease_Th",
            "PedalPosPro_Override",
            "PedalPosPro_th",
        ),
    ),
    SignalMapping("abort_any_active_event", "settingsRequest/AEB/abortAnyActiveEvents"),
    SignalMapping(
        "aeb_deceleration_request",
        "ndas_di_status/activeSafety/outputs/AEB/accelerationRequest",
    ),
)

MOTION_LOGICAL_NAMES = (
    "steering_wheel_angle",
    "steering_wheel_angle_rate",
    "yaw_rate",
    "lateral_acceleration",
)
REQUIRED_SIGNALS = (
    "vehicle_speed",
    *MOTION_LOGICAL_NAMES,
    "abort_any_active_event",
)
DEFAULT_BY_LOGICAL = {row.logical_name: row for row in DEFAULT_SIGNAL_ROWS}


def _token(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


ACRONYM_ALIASES: dict[str, str] = {
    "vehiclespeed": "vehicle_speed",
    "steeringwheelangle": "steering_wheel_angle",
    "steeringwheelanglerate": "steering_wheel_angle_rate",
    "steeringwheelanglespeed": "steering_wheel_angle_rate",
    "yawrate": "yaw_rate",
    "yawratesuspension": "yaw_rate",
    "lateralacceleration": "lateral_acceleration",
    "pedalpospro": "throttle",
    "throttle": "throttle",
    "abortanyactiveevent": "abort_any_active_event",
    "abortanyactiveevents": "abort_any_active_event",
    "aebdecelerationrequest": "aeb_deceleration_request",
    "accelerationrequest": "aeb_deceleration_request",
    "throttleincrease": "throttle_increase",
    "maxthrottle": "max_throttle",
    "throttleoverride": "throttle_override",
}

HEADER_ALIASES = {
    "acronym": {"acronym", "logicalname", "softwareacronym"},
    "model_logger": {"modellogger", "mdfsignal", "signal", "channel"},
    "cal_thd": {"calthd", "calibrationthreshold", "threshold", "calibration"},
}

CAL_PARAM_HEADER_ALIASES = {
    "parameter": {"parameter", "parametername", "calibration", "calibratable"},
    "x": {"x", "xentry", "xparameter", "xname"},
    "y": {"y", "yentry", "yparameter", "yname"},
}

MODEL_LOGGER_ALIASES = {
    ("lateral_acceleration", "rovlateralaccceleration"): "rov/lateralAcceleration",
    ("yaw_rate", "rovyawratesuppression"): "rov/YawrateSuspension",
}


def _split_calibrations(value: object) -> tuple[str, ...]:
    if value is None:
        return ()
    return tuple(part.strip() for part in re.split(r"[;,\n]", str(value)) if part.strip())


def default_mapping(source: str = "Built-in model-logger mapping") -> MappingConfiguration:
    warning = (
        "The workbook does not contain an 'swIntfc' worksheet. "
        "The built-in mapping for the supplied model-logger signals is being used."
    )
    return MappingConfiguration(
        signals={row.logical_name: row for row in DEFAULT_SIGNAL_ROWS},
        source=source,
        warning=warning,
    )


def load_mapping(path: str | Path) -> MappingConfiguration:
    workbook_path = Path(path)
    values = _read_sheet_rows(workbook_path, "swIntfc")
    if values is None:
        return default_mapping(source=f"Built-in fallback ({workbook_path.name})")
    header_index, columns = _find_headers(values, workbook_path.name)
    rows: dict[str, SignalMapping] = {}
    corrections: list[str] = []
    for row_values in values[header_index + 1 :]:
        acronym = _row_value(row_values, columns["acronym"])
        model_logger = _row_value(row_values, columns["model_logger"])
        cal_thd = _row_value(row_values, columns["cal_thd"])
        if acronym is None and model_logger is None and cal_thd is None:
            continue
        logical_name = ACRONYM_ALIASES.get(_token(acronym))
        if logical_name is None:
            continue
        configured_signal = str(model_logger or "").strip()
        corrected_signal = MODEL_LOGGER_ALIASES.get(
            (logical_name, _token(configured_signal)),
            configured_signal,
        )
        if corrected_signal != configured_signal:
            corrections.append(f"{configured_signal} -> {corrected_signal}")
        rows[logical_name] = SignalMapping(
            logical_name=logical_name,
            model_logger=corrected_signal,
            calibrations=_split_calibrations(cal_thd),
        )
    errors: list[str] = []
    for logical_name in REQUIRED_SIGNALS:
        row = rows.get(logical_name)
        if row is None:
            errors.append(f"missing acronym for {logical_name}")
        elif not row.model_logger:
            errors.append(f"modelLogger is blank for {logical_name}")
    for logical_name in MOTION_LOGICAL_NAMES:
        row = rows.get(logical_name)
        if row is not None and not row.calibrations:
            errors.append(f"cal_thd is blank for {logical_name}")
    if errors:
        raise MappingError(
            f"The 'swIntfc' worksheet in {workbook_path.name} is invalid:\n- "
            + "\n- ".join(errors)
        )
    warning = None
    if corrections:
        warning = "Corrected known modelLogger aliases: " + "; ".join(corrections)
    return MappingConfiguration(
        signals=rows,
        source=f"{workbook_path.name}:swIntfc",
        warning=warning,
    )


def load_calibration_specs(path: str | Path) -> tuple[CalibrationBindingSpec, ...]:
    workbook_path = Path(path)
    values = _read_sheet_rows(workbook_path, "calParam")
    if values is None:
        raise MappingError(f"{workbook_path.name} does not contain a 'calParam' worksheet")
    header_index, columns = _find_cal_param_headers(values, workbook_path.name)
    specs: list[CalibrationBindingSpec] = []
    errors: list[str] = []
    seen: set[str] = set()
    for row_number, row_values in enumerate(values[header_index + 1 :], start=header_index + 2):
        parameter = str(_row_value(row_values, columns["parameter"]) or "").strip()
        x_entry = str(_row_value(row_values, columns["x"]) or "").strip()
        y_entry = str(_row_value(row_values, columns["y"]) or "").strip()
        if not parameter and not x_entry and not y_entry:
            continue
        if not parameter or not x_entry or not y_entry:
            errors.append(f"row {row_number} must contain a parameter, x entry, and y entry")
            continue
        token = _token(parameter)
        if token in seen:
            errors.append(f"row {row_number} duplicates parameter {parameter!r}")
            continue
        seen.add(token)
        specs.append(
            CalibrationBindingSpec(
                parameter_name=parameter,
                x_entry_name=x_entry,
                y_entry_name=y_entry,
                source=f"{workbook_path.name}:calParam row {row_number}",
            )
        )
    if errors:
        raise MappingError(
            f"The 'calParam' worksheet in {workbook_path.name} is invalid:\n- "
            + "\n- ".join(errors)
        )
    if not specs:
        raise MappingError(f"The 'calParam' worksheet in {workbook_path.name} contains no parameter rows")
    return tuple(specs)


def match_motion_calibration_specs(
    mapping: MappingConfiguration,
    specs: tuple[CalibrationBindingSpec, ...],
) -> dict[str, CalibrationBindingSpec]:
    by_parameter = {_token(spec.parameter_name): spec for spec in specs}
    matched: dict[str, CalibrationBindingSpec] = {}
    missing: list[str] = []
    for logical_name in MOTION_LOGICAL_NAMES:
        requested_name = mapping.signal(logical_name).calibrations[0]
        spec = by_parameter.get(_token(requested_name))
        if spec is None:
            missing.append(requested_name)
        else:
            matched[logical_name] = spec
    if missing:
        raise MappingError(
            "The 'calParam' worksheet does not define the required motion parameters:\n- "
            + "\n- ".join(missing)
        )
    return matched


def _find_headers(values: list[list[object]], workbook_name: str) -> tuple[int, dict[str, int]]:
    for row_index, row_values in enumerate(values[:30]):
        columns: dict[str, int] = {}
        for index, value in enumerate(row_values):
            normalised = _token(value)
            for logical_header, aliases in HEADER_ALIASES.items():
                if normalised in aliases:
                    columns[logical_header] = index
        if len(columns) == len(HEADER_ALIASES):
            return row_index, columns
    raise MappingError(
        f"The 'swIntfc' worksheet in {workbook_name} "
        "must contain acronym, modelLogger, and cal_thd columns"
    )


def _find_cal_param_headers(
    values: list[list[object]],
    workbook_name: str,
) -> tuple[int, dict[str, int]]:
    for row_index, row_values in enumerate(values[:30]):
        columns: dict[str, int] = {}
        for index, value in enumerate(row_values):
            normalised = _token(value)
            for logical_header, aliases in CAL_PARAM_HEADER_ALIASES.items():
                if normalised in aliases:
                    columns[logical_header] = index
        if "x" in columns and "y" in columns:
            if "parameter" not in columns:
                inferred = min(columns["x"], columns["y"]) - 1
                if inferred < 0:
                    continue
                columns["parameter"] = inferred
            return row_index, columns
    raise MappingError(
        f"The 'calParam' worksheet in {workbook_name} must contain x and y columns "
        "with the parameter name in the preceding column"
    )


def _row_value(values: list[object] | tuple[object, ...], index: int) -> object:
    return values[index] if index < len(values) else None


def _read_sheet_rows(workbook_path: Path, requested_sheet: str) -> list[list[object]] | None:
    suffix = workbook_path.suffix.lower()
    if suffix not in {".xlsx", ".xlsm", ".numbers"}:
        raise MappingError(
            f"Configuration workbook must be an .xlsx, .xlsm, or .numbers file: {workbook_path}"
        )
    if suffix == ".numbers":
        try:
            from numbers_parser import Document
        except ImportError as exc:
            raise MappingError(
                "Reading .numbers files requires the numbers-parser package"
            ) from exc
        try:
            with warnings.catch_warnings():
                warnings.filterwarnings("ignore", message="unsupported version.*", category=RuntimeWarning)
                document = Document(str(workbook_path))
            sheet = next(
                (sheet for sheet in document.sheets if sheet.name.lower() == requested_sheet.lower()),
                None,
            )
            if sheet is None:
                return None
            if not sheet.tables:
                return []
            return [
                [getattr(cell, "value", None) for cell in row]
                for row in sheet.tables[0].rows()
            ]
        except MappingError:
            raise
        except Exception as exc:
            raise MappingError(f"Cannot read Numbers workbook {workbook_path}: {exc}") from exc

    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except (OSError, ValueError) as exc:
        raise MappingError(f"Cannot read configuration workbook {workbook_path}: {exc}") from exc
    try:
        sheet_name = next(
            (name for name in workbook.sheetnames if name.lower() == requested_sheet.lower()),
            None,
        )
        if sheet_name is None:
            return None
        return [list(row) for row in workbook[sheet_name].iter_rows(values_only=True)]
    finally:
        workbook.close()


def create_mapping_template(path: str | Path) -> Path:
    output = Path(path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "swIntfc"
    sheet.append(["acronym", "modelLogger", "cal_thd"])
    acronym_labels = {
        "vehicle_speed": "vehicleSpeed",
        "steering_wheel_angle": "steeringWheelAngle",
        "steering_wheel_angle_rate": "steeringWheelAngleRate",
        "yaw_rate": "yawRate",
        "lateral_acceleration": "lateralAcceleration",
        "throttle": "PedalPosPro",
        "abort_any_active_event": "abort_any_active_event",
        "aeb_deceleration_request": "aeb_deceleration_request",
    }
    for row in DEFAULT_SIGNAL_ROWS:
        sheet.append([acronym_labels[row.logical_name], row.model_logger, "; ".join(row.calibrations)])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:C{sheet.max_row}"
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 72
    sheet.column_dimensions["C"].width = 38
    parameters = workbook.create_sheet("calParam")
    parameters.append(["parameter", "x", "y"])
    for logical_name in MOTION_LOGICAL_NAMES:
        parameter_name = DEFAULT_BY_LOGICAL[logical_name].calibrations[0]
        parameters.append([parameter_name, f"{parameter_name}_x", f"{parameter_name}_y"])
    parameters.freeze_panes = "A2"
    parameters.auto_filter.ref = f"A1:C{parameters.max_row}"
    parameters.column_dimensions["A"].width = 38
    parameters.column_dimensions["B"].width = 44
    parameters.column_dimensions["C"].width = 44
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output
